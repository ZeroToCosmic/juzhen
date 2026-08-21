"""Read-only Excel parsing and validation for imported comment trees."""

from __future__ import annotations

from collections.abc import Callable, Mapping
from io import BytesIO
from itertools import islice
from pathlib import Path
from typing import Any
from uuid import uuid4
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from .errors import CampaignValidationError


MAX_IMPORT_BYTES = 2 * 1024 * 1024
MAX_IMPORT_ROWS = 5000
MAX_TREE_NODES = 100
MAX_COMMENT_LENGTH = 2200
MAX_IMPORT_WORKSHEETS = 8
MAX_IMPORT_COLUMNS = 64
MAX_IMPORT_ARCHIVE_MEMBERS = 128
MAX_IMPORT_MEMBER_UNCOMPRESSED_BYTES = 16 * 1024 * 1024
MAX_IMPORT_UNCOMPRESSED_BYTES = 24 * 1024 * 1024

HEADER_ALIASES = {
    "tree_name": {"评论树名称", "tree_name"},
    "node_no": {"节点序号", "node_no"},
    "parent_node_no": {"回复节点序号", "parent_node_no"},
    "text": {"评论文案", "comment_text", "text"},
}


def preview_comment_tree_workbook(filename: str, content: bytes) -> dict[str, Any]:
    """Return normalized, tree-scoped preview data without persisting anything."""
    if Path(filename or "").suffix.lower() != ".xlsx":
        raise CampaignValidationError("unsupported_import_type")
    if not content:
        raise CampaignValidationError("import_file_invalid")
    if len(content) > MAX_IMPORT_BYTES:
        raise CampaignValidationError("import_file_too_large")
    _validate_archive(content)
    workbook = None
    try:
        workbook = load_workbook(
            BytesIO(content), read_only=True, data_only=True, keep_links=False
        )
        if len(workbook.worksheets) > MAX_IMPORT_WORKSHEETS:
            raise CampaignValidationError("import_file_invalid")
        worksheet = workbook.worksheets[0]
        if (worksheet.max_row or 0) > MAX_IMPORT_ROWS + 1:
            raise CampaignValidationError("import_file_too_large")
        if (worksheet.max_column or 0) > MAX_IMPORT_COLUMNS:
            raise CampaignValidationError("import_file_invalid")
        rows = list(
            islice(
                worksheet.iter_rows(
                    max_row=MAX_IMPORT_ROWS + 1,
                    max_col=MAX_IMPORT_COLUMNS,
                    values_only=True,
                ),
                MAX_IMPORT_ROWS + 2,
            )
        )
    except CampaignValidationError:
        raise
    except Exception as exc:
        raise CampaignValidationError("import_file_invalid") from exc
    finally:
        if workbook is not None:
            workbook.close()

    if len(rows) > MAX_IMPORT_ROWS + 1:
        raise CampaignValidationError("import_file_too_large")
    headers = _map_headers(rows[0] if rows else ())
    grouped = _normalize_rows(rows[1:], headers)
    trees = [
        normalize_imported_tree({"name": name, "nodes": nodes}, source_locations=True)
        for name, nodes in grouped.items()
    ]
    return {
        "trees": trees,
        "summary": {
            "tree_count": len(trees),
            "valid_count": sum(tree["valid"] for tree in trees),
            "rejected_count": sum(not tree["valid"] for tree in trees),
        },
    }


def import_tree_to_template(
    tree: Mapping[str, Any], *, id_factory: Callable[[], str] = lambda: str(uuid4())
) -> dict[str, Any]:
    """Convert one validated preview tree into the existing TemplateCreate shape."""
    checked = normalize_imported_tree(tree)
    if not checked["valid"]:
        raise CampaignValidationError("template_invalid")
    id_by_no = {node["node_no"]: str(id_factory()) for node in checked["nodes"]}
    if len(id_by_no) != len(set(id_by_no.values())) or any(not value for value in id_by_no.values()):
        raise CampaignValidationError("template_invalid")
    steps = []
    for position, node in enumerate(checked["nodes"]):
        parent_no = node["parent_node_no"]
        steps.append(
            {
                "id": id_by_no[node["node_no"]],
                "label": "楼主评论" if parent_no is None else f"回复 {position}",
                "content_source": "fixed",
                "fixed_text": node["text"],
                "content_library_id": "",
                "content_item_id": "",
                "parent_step_id": id_by_no[parent_no] if parent_no else None,
                "required_profile_tags": [],
                "excluded_profile_tags": [],
                "language": "",
            }
        )
    return {
        "name": checked["name"],
        "description": "",
        "supported_modes": ["threaded"],
        "language": "",
        "tags": [],
        "steps": steps,
    }


def _map_headers(values: tuple[Any, ...] | list[Any]) -> dict[str, int]:
    mapped: dict[str, int] = {}
    for index, value in enumerate(values):
        normalized = _text(value).casefold()
        for field, aliases in HEADER_ALIASES.items():
            if normalized in aliases:
                if field in mapped:
                    raise CampaignValidationError("import_file_invalid")
                mapped[field] = index
    if len(mapped) != len(HEADER_ALIASES):
        raise CampaignValidationError("import_file_invalid")
    return mapped


def _normalize_rows(
    rows: list[tuple[Any, ...]], headers: Mapping[str, int]
) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row_number, values in enumerate(rows, start=2):
        if not any(_text(value) for value in values):
            continue
        name = _cell(values, headers, "tree_name")
        nodes = grouped.setdefault(name, [])
        nodes.append(
            {
                "node_no": _cell(values, headers, "node_no"),
                "parent_node_no": _cell(values, headers, "parent_node_no") or None,
                "text": _cell(values, headers, "text"),
                "row": row_number,
                "position": len(nodes),
            }
        )
    return grouped


def normalize_imported_tree(
    tree: Mapping[str, Any], *, source_locations: bool = False
) -> dict[str, Any]:
    """Normalize raw tree input and return row-aware, tree-scoped validation data.

    ``source_locations`` is reserved for workbook rows.  Client commit input
    must use the default so preview-provided validity and display locations
    cannot affect the server-side conversion.
    """
    normalized_name = _text(tree.get("name"))
    raw_nodes = tree.get("nodes")
    normalized_nodes = _normalize_nodes(
        list(raw_nodes) if isinstance(raw_nodes, list) else [],
        source_locations=source_locations,
    )
    fallback_row = normalized_nodes[0]["row"] if normalized_nodes else 2
    errors: list[dict[str, Any]] = []

    if not normalized_name:
        _error(errors, "tree_name_missing", fallback_row)
    elif len(normalized_name) > 100:
        _error(errors, "tree_name_invalid", fallback_row)
    if not normalized_nodes:
        _error(errors, "tree_empty", fallback_row)
    if len(normalized_nodes) > MAX_TREE_NODES:
        _error(errors, "tree_too_large", normalized_nodes[MAX_TREE_NODES]["row"])

    first_by_node_no: dict[str, dict[str, Any]] = {}
    valid_nodes: list[dict[str, Any]] = []
    for node in normalized_nodes:
        node_no = node["node_no"]
        if not node_no:
            _error(errors, "node_no_missing", node["row"])
        elif len(node_no) > 120:
            _error(errors, "node_no_invalid", node["row"])
        elif node_no in first_by_node_no:
            _error(errors, "duplicate_node_no", node["row"])
        else:
            first_by_node_no[node_no] = node
            valid_nodes.append(node)

        if not node["text"]:
            _error(errors, "comment_text_missing", node["row"])
        elif len(node["text"]) > MAX_COMMENT_LENGTH:
            _error(errors, "comment_text_too_long", node["row"])

    parents = {node["node_no"]: node["parent_node_no"] for node in valid_nodes}
    for node in valid_nodes:
        parent = node["parent_node_no"]
        if parent is not None and parent not in parents:
            _error(errors, "parent_not_found", node["row"])

    roots = [node for node in valid_nodes if node["parent_node_no"] is None]
    if len(roots) != 1:
        _error(errors, "root_count_invalid", fallback_row)

    _detect_cycles(parents, first_by_node_no, errors)
    return {
        "name": normalized_name,
        "nodes": normalized_nodes,
        "errors": errors,
        "valid": not errors,
    }


def _normalize_nodes(
    nodes: list[Any], *, source_locations: bool
) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for index, raw in enumerate(nodes):
        raw = raw if isinstance(raw, Mapping) else {}
        node_no = _text(raw.get("node_no"))
        parent_node_no = _text(raw.get("parent_node_no")) or None
        if parent_node_no is None and normalized:
            parent_node_no = normalized[-1]["node_no"] or None
        row = raw.get("row") if source_locations else index + 2
        position = raw.get("position") if source_locations else index
        normalized.append(
            {
                "node_no": node_no,
                "parent_node_no": parent_node_no,
                "text": _text(raw.get("text")),
                "row": row if isinstance(row, int) and row >= 2 else index + 2,
                "position": position if isinstance(position, int) and position >= 0 else index,
            }
        )
    return normalized


def _detect_cycles(
    parents: Mapping[str, str | None],
    nodes: Mapping[str, Mapping[str, Any]],
    errors: list[dict[str, Any]],
) -> None:
    colors: dict[str, int] = {}
    reported: set[str] = set()

    def visit(node_no: str) -> None:
        color = colors.get(node_no, 0)
        if color == 1:
            if node_no not in reported:
                _error(errors, "cycle_detected", nodes[node_no]["row"])
                reported.add(node_no)
            return
        if color == 2:
            return
        colors[node_no] = 1
        parent = parents[node_no]
        if parent in parents:
            visit(parent)
        colors[node_no] = 2

    for node_no in parents:
        visit(node_no)


def _cell(values: tuple[Any, ...], headers: Mapping[str, int], field: str) -> str:
    index = headers[field]
    return _text(values[index]) if index < len(values) else ""


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _error(errors: list[dict[str, Any]], code: str, row: int) -> None:
    errors.append({"code": code, "row": row})


def _validate_archive(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
    except (BadZipFile, OSError, RuntimeError) as exc:
        raise CampaignValidationError("import_file_invalid") from exc
    if not members or len(members) > MAX_IMPORT_ARCHIVE_MEMBERS:
        raise CampaignValidationError("import_file_invalid")

    total_uncompressed = 0
    for member in members:
        if member.flag_bits & 0x1:
            raise CampaignValidationError("import_file_invalid")
        if member.file_size > MAX_IMPORT_MEMBER_UNCOMPRESSED_BYTES:
            raise CampaignValidationError("import_file_invalid")
        total_uncompressed += member.file_size
        if total_uncompressed > MAX_IMPORT_UNCOMPRESSED_BYTES:
            raise CampaignValidationError("import_file_invalid")

from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook
from sqlalchemy.exc import IntegrityError

from comment_campaign.errors import CampaignValidationError
from comment_campaign.schemas import TemplateImportCommit
from comment_campaign.service import CommentCampaignService
from comment_campaign.store import CampaignStore
from comment_campaign.template_import import (
    MAX_COMMENT_LENGTH,
    MAX_IMPORT_ARCHIVE_MEMBERS,
    MAX_IMPORT_BYTES,
    MAX_IMPORT_COLUMNS,
    MAX_IMPORT_MEMBER_UNCOMPRESSED_BYTES,
    MAX_IMPORT_ROWS,
    MAX_IMPORT_WORKSHEETS,
    MAX_TREE_NODES,
    import_tree_to_template,
    normalize_imported_tree,
    preview_comment_tree_workbook,
)


@pytest.fixture
def xlsx_bytes():
    def build(rows):
        workbook = Workbook()
        sheet = workbook.active
        for row in rows:
            sheet.append(row)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    return build


def _sheet(*rows):
    return [
        ["评论树名称", "节点序号", "回复节点序号", "评论文案"],
        *rows,
    ]


def _workbook_bytes(workbook):
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _encrypted_zip_bytes():
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        archive.writestr("sheet.xml", "not important")
    encrypted = bytearray(output.getvalue())
    for signature, flag_offset in ((b"PK\x03\x04", 6), (b"PK\x01\x02", 8)):
        start = 0
        while (offset := encrypted.find(signature, start)) >= 0:
            encrypted[offset + flag_offset] |= 0x1
            start = offset + len(signature)
    return bytes(encrypted)


@pytest.fixture
def import_service(tmp_path):
    store = CampaignStore(f"sqlite:///{tmp_path / 'imports.db'}")
    store.initialize()
    return CommentCampaignService(store)


def test_preview_groups_linear_and_branched_trees(xlsx_bytes):
    result = preview_comment_tree_workbook("trees.xlsx", xlsx_bytes(_sheet(
        ["A", "1", "", "root"],
        ["A", "2", "", "linear child"],
        ["A", "3", "1", "branch child"],
        ["B", "1", "", "other root"],
    )))

    assert [tree["name"] for tree in result["trees"]] == ["A", "B"]
    assert result["trees"][0]["nodes"][1]["parent_node_no"] == "1"
    assert result["trees"][0]["nodes"][2]["parent_node_no"] == "1"
    assert all(tree["valid"] for tree in result["trees"])
    assert result["summary"] == {"tree_count": 2, "valid_count": 2, "rejected_count": 0}


def test_preview_accepts_english_headers_and_groups_non_contiguous_tree_rows(xlsx_bytes):
    result = preview_comment_tree_workbook("trees.xlsx", xlsx_bytes([
        ["tree_name", "node_no", "parent_node_no", "comment_text"],
        ["A", "1", "", "root"],
        ["B", "1", "", "other root"],
        ["A", "2", "", "later linear child"],
    ]))

    assert [node["parent_node_no"] for node in result["trees"][0]["nodes"]] == [None, "1"]
    assert all(tree["valid"] for tree in result["trees"])


def test_preview_reads_first_worksheet_even_when_another_sheet_is_active(xlsx_bytes):
    workbook = Workbook()
    first = workbook.worksheets[0]
    for row in _sheet(["A", "1", "", "root"]):
        first.append(row)
    workbook.create_sheet("active-but-ignored")
    workbook.active = 1

    result = preview_comment_tree_workbook("trees.xlsx", _workbook_bytes(workbook))

    assert result["trees"][0]["name"] == "A"


@pytest.mark.parametrize(("rows", "code"), [
    ([["A", "1", "", ""]], "comment_text_missing"),
    ([["A", "1", "", "root"], ["A", "1", "", "duplicate"]], "duplicate_node_no"),
    ([["A", "1", "2", "one"], ["A", "2", "1", "two"]], "cycle_detected"),
    ([["A", "1", "9", "orphan"]], "parent_not_found"),
    ([["A", "1", "", "root"], ["A", "2", "", "x" * (MAX_COMMENT_LENGTH + 1)]], "comment_text_too_long"),
])
def test_preview_reports_tree_scoped_errors(xlsx_bytes, rows, code):
    result = preview_comment_tree_workbook("trees.xlsx", xlsx_bytes(_sheet(*rows)))

    assert result["trees"][0]["valid"] is False
    assert code in {error["code"] for error in result["trees"][0]["errors"]}
    assert all("row" in error for error in result["trees"][0]["errors"])


def test_preview_rejects_more_than_one_hundred_nodes_per_tree(xlsx_bytes):
    rows = [["A", str(index), "", f"comment {index}"] for index in range(1, MAX_TREE_NODES + 2)]
    result = preview_comment_tree_workbook("trees.xlsx", xlsx_bytes(_sheet(*rows)))

    assert result["trees"][0]["valid"] is False
    assert {error["code"] for error in result["trees"][0]["errors"]} >= {"tree_too_large"}


def test_preview_rejects_missing_headers_invalid_files_and_row_limit(xlsx_bytes):
    with pytest.raises(CampaignValidationError):
        preview_comment_tree_workbook("trees.xlsx", xlsx_bytes([["tree_name", "node_no"], ["A", "1"]]))
    with pytest.raises(CampaignValidationError):
        preview_comment_tree_workbook("trees.xlsx", xlsx_bytes([
            ["评论树名称", "tree_name", "节点序号", "回复节点序号", "评论文案"],
            ["A", "A", "1", "", "root"],
        ]))
    with pytest.raises(CampaignValidationError):
        preview_comment_tree_workbook("trees.csv", b"x")
    with pytest.raises(CampaignValidationError):
        preview_comment_tree_workbook("trees.xlsx", b"broken")
    with pytest.raises(CampaignValidationError):
        preview_comment_tree_workbook("trees.xlsx", b"x" * (MAX_IMPORT_BYTES + 1))

    rows = [["A", str(index), "", "root"] for index in range(1, MAX_IMPORT_ROWS + 2)]
    with pytest.raises(CampaignValidationError):
        preview_comment_tree_workbook("trees.xlsx", xlsx_bytes(_sheet(*rows)))


def test_preview_rejects_unsafe_archives_before_openpyxl():
    many_members = BytesIO()
    with ZipFile(many_members, "w", ZIP_DEFLATED) as archive:
        for index in range(MAX_IMPORT_ARCHIVE_MEMBERS + 1):
            archive.writestr(f"member-{index}.xml", "x")
    oversized_member = BytesIO()
    with ZipFile(oversized_member, "w", ZIP_DEFLATED) as archive:
        archive.writestr("member.xml", b"x" * (MAX_IMPORT_MEMBER_UNCOMPRESSED_BYTES + 1))

    for content in (many_members.getvalue(), oversized_member.getvalue(), _encrypted_zip_bytes()):
        with pytest.raises(CampaignValidationError):
            preview_comment_tree_workbook("trees.xlsx", content)


def test_preview_rejects_too_many_sheets_and_columns(xlsx_bytes):
    workbook = Workbook()
    for index in range(MAX_IMPORT_WORKSHEETS):
        workbook.create_sheet(f"extra-{index}")
    with pytest.raises(CampaignValidationError):
        preview_comment_tree_workbook("trees.xlsx", _workbook_bytes(workbook))

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(_sheet()[0])
    sheet.cell(1, MAX_IMPORT_COLUMNS + 1, "overflow")
    with pytest.raises(CampaignValidationError):
        preview_comment_tree_workbook("trees.xlsx", _workbook_bytes(workbook))


def test_preview_preserves_zero_and_normalizes_numeric_node_numbers(xlsx_bytes):
    result = preview_comment_tree_workbook("trees.xlsx", xlsx_bytes(_sheet(
        ["A", 0, "", "root"],
        ["A", 1.0, 0.0, "child"],
    )))

    assert result["trees"][0]["nodes"] == [
        {"node_no": "0", "parent_node_no": None, "text": "root", "row": 2, "position": 0},
        {"node_no": "1", "parent_node_no": "0", "text": "child", "row": 3, "position": 1},
    ]


def test_preview_rejects_tree_names_over_one_hundred_characters(xlsx_bytes):
    result = preview_comment_tree_workbook("trees.xlsx", xlsx_bytes(_sheet(
        ["A" * 101, "1", "", "root"],
    )))

    assert result["trees"][0]["valid"] is False
    assert {error["code"] for error in result["trees"][0]["errors"]} >= {"tree_name_invalid"}


def test_converter_revalidates_and_generates_opaque_step_identifiers():
    identifiers = iter(["opaque-root", "opaque-child"])
    template = import_tree_to_template(
        {
            "name": "Tree",
            "nodes": [
                {"node_no": "root", "parent_node_no": None, "text": "first", "row": 2, "position": 0},
                {"node_no": "child", "parent_node_no": "root", "text": "second", "row": 3, "position": 1},
            ],
        },
        id_factory=lambda: next(identifiers),
    )

    assert template["name"] == "Tree"
    assert template["supported_modes"] == ["threaded"]
    assert [step["id"] for step in template["steps"]] == ["opaque-root", "opaque-child"]
    assert template["steps"][1]["parent_step_id"] == "opaque-root"
    assert all(step["content_source"] == "fixed" for step in template["steps"])


def test_converter_rejects_invalid_tree():
    with pytest.raises(CampaignValidationError):
        import_tree_to_template({
            "name": "broken",
            "nodes": [{"node_no": "1", "parent_node_no": "missing", "text": "text", "row": 2, "position": 0}],
        })


def test_converter_rejects_identifier_factory_collisions():
    with pytest.raises(CampaignValidationError):
        import_tree_to_template(
            {
                "name": "Tree",
                "nodes": [
                    {"node_no": "1", "parent_node_no": None, "text": "root"},
                    {"node_no": "2", "parent_node_no": "1", "text": "child"},
                ],
            },
            id_factory=lambda: "same-id",
        )


def test_public_normalizer_and_converter_ignore_tampered_preview_fields():
    raw = {
        "name": "Tree",
        "valid": True,
        "errors": [],
        "nodes": [
            {"node_no": "root", "parent_node_no": None, "text": "first", "row": 999, "position": 99},
            {"node_no": "child", "parent_node_no": "root", "text": "second", "row": 998, "position": 98},
        ],
    }
    normalized = normalize_imported_tree(raw)

    assert normalized["valid"] is True
    assert [(node["row"], node["position"]) for node in normalized["nodes"]] == [(2, 0), (3, 1)]

    raw["nodes"][1]["parent_node_no"] = "missing"
    with pytest.raises(CampaignValidationError):
        import_tree_to_template(raw)


def _tree(name, nodes):
    return {"name": name, "nodes": nodes}


def _node(node_no, text, parent_node_no=None):
    return {"node_no": node_no, "parent_node_no": parent_node_no, "text": text}


def test_commit_schema_rejects_derived_fields_but_defers_business_limits():
    with pytest.raises(Exception):
        TemplateImportCommit.model_validate({
            "trees": [{"name": "A", "valid": True, "nodes": [_node("1", "root")]}],
        })
    with pytest.raises(Exception):
        TemplateImportCommit.model_validate({
            "trees": [{
                "name": "A",
                "nodes": [{**_node("1", "root"), "row": 2}],
            }],
        })

    request = TemplateImportCommit.model_validate({
        "trees": [_tree("A" * 101, [_node("", "x" * (MAX_COMMENT_LENGTH + 1))])],
    })
    assert request.trees[0].name == "A" * 101


@pytest.mark.parametrize("code", [
    "unsupported_import_type",
    "import_file_invalid",
    "import_file_too_large",
    "import_tree_failed",
])
def test_import_error_codes_remain_exact(code):
    assert CampaignValidationError(code).code == code


def test_commit_schema_rejects_more_than_five_thousand_total_nodes():
    with pytest.raises(Exception):
        TemplateImportCommit.model_validate({
            "trees": [_tree("A", [_node(str(index), "text") for index in range(5001)])],
        })


def test_import_service_creates_valid_trees_and_rejects_business_invalid_trees(import_service):
    oversized_nodes = [_node(str(index), f"text {index}") for index in range(101)]
    result = import_service.import_templates({"trees": [
        _tree("A", [_node("1", "root"), _node("2", "child", "1")]),
        _tree("B", oversized_nodes),
        _tree("C", [_node("root", "another root")]),
    ]})

    assert [item["name"] for item in result["created"]] == ["A", "C"]
    assert result["rejected"][0]["name"] == "B"
    assert {error["code"] for error in result["rejected"][0]["errors"]} >= {"tree_too_large"}
    assert import_service.store.get_template(result["created"][0]["id"])["steps"][0]["id"] != "1"


def test_import_service_continues_after_one_tree_integrity_error(tmp_path):
    class FailingService(CommentCampaignService):
        def create_template(self, payload, template_id=None):
            if payload.name == "B":
                raise IntegrityError("insert", {}, RuntimeError("duplicate"))
            return super().create_template(payload, template_id)

    store = CampaignStore(f"sqlite:///{tmp_path / 'imports.db'}")
    store.initialize()
    service = FailingService(store)
    result = service.import_templates({"trees": [
        _tree("A", [_node("1", "root")]),
        _tree("B", [_node("1", "root")]),
        _tree("C", [_node("1", "root")]),
    ]})

    assert [item["name"] for item in result["created"]] == ["A", "C"]
    assert result["rejected"] == [{"name": "B", "errors": [{"code": "import_tree_failed"}]}]


def test_import_service_does_not_hide_unknown_persistence_errors(tmp_path):
    class FailingService(CommentCampaignService):
        def create_template(self, _payload, _template_id=None):
            raise RuntimeError("database secret")

    store = CampaignStore(f"sqlite:///{tmp_path / 'imports.db'}")
    store.initialize()
    service = FailingService(store)

    with pytest.raises(RuntimeError, match="database secret"):
        service.import_templates({"trees": [_tree("A", [_node("1", "root")])]})

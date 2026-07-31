import csv
from io import TextIOWrapper
from pathlib import Path

from openpyxl import load_workbook


HEADER_ALIASES = {
    "brand_name": {"品牌名", "品牌", "brand_name", "brand"},
    "body": {"文案", "正文", "copy", "body", "content"},
    "tags": {"tag", "标签", "tags"},
}


def normalize_header(value):
    return str(value or "").strip().casefold()


def map_headers(values):
    mapped = {}
    for index, value in enumerate(values):
        normalized = normalize_header(value)
        for field, aliases in HEADER_ALIASES.items():
            if normalized in aliases and field not in mapped:
                mapped[field] = index

    if "brand_name" not in mapped or "body" not in mapped:
        raise ValueError("缺少必要表头：品牌名、文案")
    return mapped


def normalize_row(row_number, values, headers):
    def cell(field):
        index = headers.get(field)
        if index is None or index >= len(values):
            return ""
        return str(values[index] or "").strip()

    brand_name = cell("brand_name")
    body = cell("body")
    if not brand_name:
        return None, {"row": row_number, "error": "缺少品牌名"}
    if not body:
        return None, {"row": row_number, "error": "缺少文案"}
    return (
        {
            "row": row_number,
            "brand_name": brand_name,
            "body": body,
            "tags": cell("tags"),
        },
        None,
    )


def parse_rows(rows):
    rows = iter(rows)
    try:
        headers = map_headers(next(rows))
    except StopIteration as error:
        raise ValueError("导入文件为空") from error

    parsed = []
    errors = []
    total = 0
    for row_number, values in enumerate(rows, start=2):
        values = list(values)
        if not any(str(value or "").strip() for value in values):
            continue
        total += 1
        item, item_error = normalize_row(row_number, values, headers)
        if item_error:
            errors.append(item_error)
        else:
            parsed.append(item)
    return {"total": total, "rows": parsed, "errors": errors}


def parse_copy_import(filename, stream):
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".xlsx":
        try:
            workbook = load_workbook(stream, read_only=True, data_only=True)
            return parse_rows(workbook.active.iter_rows(values_only=True))
        except ValueError:
            raise
        except Exception as error:
            raise ValueError("Excel 文件无法读取") from error

    if suffix in {".csv", ".tsv"}:
        try:
            wrapper = TextIOWrapper(stream, encoding="utf-8-sig", newline="")
            delimiter = "," if suffix == ".csv" else "\t"
            return parse_rows(csv.reader(wrapper, delimiter=delimiter))
        except UnicodeError as error:
            raise ValueError("表格文件必须使用 UTF-8 编码") from error

    raise ValueError("仅支持 .xlsx、.csv、.tsv 文件")
